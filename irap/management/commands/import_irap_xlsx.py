from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from irap.models import (
    Boundary,
    BoundaryAssociation,
    Control,
    ControlAssessment,
    ControlSet,
    Evidence,
    EvidenceLink,
    EvidenceUsage,
)

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - handled at runtime
    load_workbook = None


def _canon(text: str) -> str:
    return "".join(ch for ch in str(text).strip().lower() if ch.isalnum())


def _as_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _status_value(raw: str) -> str:
    v = _canon(raw)
    mapping = {
        "pass": ControlAssessment.Status.PASS_,
        "passed": ControlAssessment.Status.PASS_,
        "partial": ControlAssessment.Status.PARTIAL,
        "partiallyimplemented": ControlAssessment.Status.PARTIAL,
        "fail": ControlAssessment.Status.FAIL,
        "failed": ControlAssessment.Status.FAIL,
        "na": ControlAssessment.Status.N_A,
        "notapplicable": ControlAssessment.Status.N_A,
        "notassessed": ControlAssessment.Status.NOT_ASSESSED,
        "unassessed": ControlAssessment.Status.NOT_ASSESSED,
    }
    return mapping.get(v, ControlAssessment.Status.NOT_ASSESSED)


def _evidence_type_value(raw: str) -> str:
    v = _canon(raw)
    mapping = {
        "policy": Evidence.EvidenceType.POLICY,
        "procedure": Evidence.EvidenceType.PROCEDURE,
        "procedures": Evidence.EvidenceType.PROCEDURE,
        "diagram": Evidence.EvidenceType.DIAGRAM,
        "architecturediagram": Evidence.EvidenceType.DIAGRAM,
        "ticket": Evidence.EvidenceType.TICKET,
        "changeticket": Evidence.EvidenceType.TICKET,
        "log": Evidence.EvidenceType.LOG,
        "logging": Evidence.EvidenceType.LOG,
        "monitoring": Evidence.EvidenceType.LOG,
        "attestation": Evidence.EvidenceType.ATTESTATION,
        "interview": Evidence.EvidenceType.ATTESTATION,
    }
    return mapping.get(v, Evidence.EvidenceType.OTHER)


def _first_value(row: Dict[str, str], aliases: Iterable[str], required: bool = False) -> str:
    for alias in aliases:
        if alias in row and _as_str(row[alias]):
            return _as_str(row[alias])
    if required:
        raise CommandError(f"Missing required column value for aliases: {', '.join(aliases)}")
    return ""


def _sheet_rows(ws) -> Iterable[Dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    canon_headers = [_canon(h) for h in headers]
    out = []
    for r in rows[1:]:
        if r is None:
            continue
        values = [_as_str(v) for v in r]
        if not any(values):
            continue
        out.append({canon_headers[i]: values[i] for i in range(len(canon_headers))})
    return out


@dataclass
class Stats:
    controls: int = 0
    evidence: int = 0
    usages: int = 0
    links: int = 0
    boundaries: int = 0
    boundary_associations: int = 0
    assessments: int = 0


class Command(BaseCommand):
    help = "Import IRAP workbook (.xlsx) into the Django IRAP data model."

    def add_arguments(self, parser):
        parser.add_argument("--xlsx", required=True, help="Absolute path to source .xlsx workbook")
        parser.add_argument("--control-set-name", default="ISM Real Import")
        parser.add_argument("--control-set-version", default="September 2025")
        parser.add_argument("--control-set-description", default="Imported from real IRAP workbook.")
        parser.add_argument("--dry-run", action="store_true", help="Validate and parse without writing")

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError(
                "openpyxl is required. Install it in your venv with: pip install openpyxl"
            )

        xlsx_path = Path(options["xlsx"]).expanduser().resolve()
        if not xlsx_path.exists():
            raise CommandError(f"Workbook not found: {xlsx_path}")

        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        sheets = { _canon(name): wb[name] for name in wb.sheetnames }

        stats = Stats()
        with transaction.atomic():
            control_set, _ = ControlSet.objects.update_or_create(
                name=options["control_set_name"],
                version=options["control_set_version"],
                defaults={"description": options["control_set_description"]},
            )

            control_index = self._import_controls(sheets, control_set, stats)
            evidence_index = self._import_evidence(sheets, stats)
            self._import_usages(sheets, control_index, evidence_index, stats)
            self._import_links(sheets, evidence_index, stats)
            boundary_index = self._import_boundaries(sheets, stats)
            self._import_boundary_associations(sheets, evidence_index, boundary_index, stats)
            self._import_assessments(sheets, control_index, stats)

            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("Dry run complete. Transaction rolled back."))
                return

        self.stdout.write(self.style.SUCCESS(f"Imported workbook: {xlsx_path}"))
        self.stdout.write(
            self.style.SUCCESS(
                "Created/updated -> "
                f"controls: {stats.controls}, evidence: {stats.evidence}, usages: {stats.usages}, "
                f"links: {stats.links}, boundaries: {stats.boundaries}, "
                f"boundary_associations: {stats.boundary_associations}, assessments: {stats.assessments}"
            )
        )

    def _pick_sheet(self, sheets, aliases: Iterable[str]):
        alias_keys = [_canon(a) for a in aliases]
        # Exact/starts-with match first.
        for alias in aliases:
            key = _canon(alias)
            for sheet_key, ws in sheets.items():
                if sheet_key == key or sheet_key.startswith(key):
                    return ws
        # Then fallback: contains alias token.
        for sheet_key, ws in sheets.items():
            if any(a in sheet_key for a in alias_keys):
                return ws
        return None

    def _import_controls(self, sheets, control_set, stats: Stats):
        ws = self._pick_sheet(sheets, ["controls", "control", "irapcontrols"])
        if ws is None:
            raise CommandError("Missing required sheet: Controls")
        control_index = {}
        for row in _sheet_rows(ws):
            control_id = _first_value(row, ["controlid", "ismid", "identifier", "id"])
            if not control_id:
                continue

            name = _first_value(row, ["name", "controlname", "title", "topic"])
            description = _first_value(row, ["description", "controldescription"])
            version = _first_value(row, ["version", "controlversion", "revision"])
            guideline = _first_value(row, ["guideline"])
            section = _first_value(row, ["section"])
            updated = _first_value(row, ["updated"])

            # Skip non-data/template rows from SSP annex files.
            if _canon(control_id) in {"identifier", "id"}:
                continue
            if _canon(guideline).startswith("ignorethisline"):
                continue

            # Some SSP annex rows may omit topic/name while still carrying Identifier.
            if not name:
                fallback_parts = [section, guideline]
                fallback = " - ".join([p for p in fallback_parts if p])
                name = fallback if fallback else f"Imported control {control_id}"

            if not description:
                parts = []
                if guideline:
                    parts.append(f"Guideline: {guideline}")
                if section:
                    parts.append(f"Section: {section}")
                if updated:
                    parts.append(f"Updated: {updated}")
                description = " | ".join(parts)

            control, _ = Control.objects.update_or_create(
                control_set=control_set,
                control_id=control_id,
                version=version,
                defaults={"name": name, "description": description},
            )
            control_index[_canon(control_id)] = control
            stats.controls += 1
        return control_index

    def _import_evidence(self, sheets, stats: Stats):
        ws = self._pick_sheet(sheets, ["evidence", "evidenceartifacts", "artifacts"])
        if ws is None:
            self.stdout.write(self.style.WARNING("No Evidence sheet detected. Importing controls-only dataset."))
            return {}
        evidence_index = {}
        for row in _sheet_rows(ws):
            name = _first_value(row, ["name", "evidencename", "artifact"], required=True)
            raw_type = _first_value(row, ["type", "evidencetype", "category"])
            description = _first_value(row, ["description", "evidencedescription"])

            evidence, _ = Evidence.objects.update_or_create(
                name=name,
                defaults={
                    "type": _evidence_type_value(raw_type),
                    "description": description,
                },
            )
            evidence_index[_canon(name)] = evidence
            stats.evidence += 1
        return evidence_index

    def _import_usages(self, sheets, control_index, evidence_index, stats: Stats):
        ws = self._pick_sheet(sheets, ["evidenceusage", "usages", "linkscontroltoevidence"])
        if ws is None:
            return
        for row in _sheet_rows(ws):
            control_id = _first_value(row, ["controlid", "ismid", "control"], required=True)
            evidence_name = _first_value(row, ["evidencename", "evidence", "artifact"], required=True)
            notes = _first_value(row, ["notes", "usagenotes", "rationale"])

            control = control_index.get(_canon(control_id))
            evidence = evidence_index.get(_canon(evidence_name))
            if not control or not evidence:
                continue

            EvidenceUsage.objects.update_or_create(
                control=control,
                evidence=evidence,
                defaults={"notes": notes},
            )
            stats.usages += 1

    def _import_links(self, sheets, evidence_index, stats: Stats):
        ws = self._pick_sheet(sheets, ["evidencelinks", "links", "references"])
        if ws is None:
            return
        for row in _sheet_rows(ws):
            evidence_name = _first_value(row, ["evidencename", "evidence", "artifact"], required=True)
            url_or_reference = _first_value(row, ["urlorreference", "reference", "url"], required=True)
            description = _first_value(row, ["description", "linkdescription"])
            evidence = evidence_index.get(_canon(evidence_name))
            if not evidence:
                continue
            EvidenceLink.objects.update_or_create(
                evidence=evidence,
                url_or_reference=url_or_reference,
                defaults={"description": description},
            )
            stats.links += 1

    def _import_boundaries(self, sheets, stats: Stats):
        ws = self._pick_sheet(sheets, ["boundaries", "boundary"])
        boundary_index = {}
        if ws is None:
            return boundary_index
        for row in _sheet_rows(ws):
            name = _first_value(row, ["name", "boundaryname"], required=True)
            description = _first_value(row, ["description", "boundarydescription"])
            boundary, _ = Boundary.objects.update_or_create(
                name=name,
                defaults={"description": description},
            )
            boundary_index[_canon(name)] = boundary
            stats.boundaries += 1
        return boundary_index

    def _import_boundary_associations(self, sheets, evidence_index, boundary_index, stats: Stats):
        ws = self._pick_sheet(
            sheets,
            ["boundaryassociation", "boundaryassociations", "evidenceboundary"],
        )
        if ws is None:
            return
        for row in _sheet_rows(ws):
            evidence_name = _first_value(row, ["evidencename", "evidence", "artifact"], required=True)
            boundary_name = _first_value(row, ["boundaryname", "boundary"], required=True)
            notes = _first_value(row, ["notes", "scope", "boundarynotes"])
            evidence = evidence_index.get(_canon(evidence_name))
            boundary = boundary_index.get(_canon(boundary_name))
            if not evidence or not boundary:
                continue
            BoundaryAssociation.objects.update_or_create(
                evidence=evidence,
                boundary=boundary,
                defaults={"notes": notes},
            )
            stats.boundary_associations += 1

    def _import_assessments(self, sheets, control_index, stats: Stats):
        ws = self._pick_sheet(sheets, ["assessments", "controlassessments", "assessment"])
        if ws is None:
            return
        for row in _sheet_rows(ws):
            control_id = _first_value(row, ["controlid", "ismid", "control"], required=True)
            status = _status_value(_first_value(row, ["status", "assessmentstatus"]))
            notes = _first_value(row, ["notes", "assessmentnotes"])
            assessment_date = _first_value(row, ["assessmentdate", "date"])

            control = control_index.get(_canon(control_id))
            if not control:
                continue

            defaults = {"status": status, "notes": notes}
            if assessment_date:
                defaults["assessment_date"] = assessment_date
                ControlAssessment.objects.update_or_create(
                    control=control,
                    assessment_date=assessment_date,
                    defaults=defaults,
                )
            else:
                ControlAssessment.objects.update_or_create(
                    control=control,
                    assessment_date=timezone.localdate(),
                    defaults=defaults,
                )
            stats.assessments += 1
